import docx
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Load the .docx file
doc = docx.Document("text.docx")

# Define the keywords in groups
pains = {
    "Group 1": ["difficult", "challenging", "complicated"],
    "Group 2": ["lengthy", "time-consuming", "tedious", "laborious", "slow", "arduous", "cumbersome", "complicated", "complex", "protracted", "excessive", "overwhelming", "burdensome", "dragging", "tiresome", "irritating", "frustrating", "cumbersome", "inefficient"],
    "Group 3": ["bugs", "glitches", "issues", "defects", "flaws", "malfunctions", "crashes", "failures", "mistakes", "problems", "errors", "inconsistencies", "hiccups", "snags", "setbacks", "technical difficulties", "unexpected behavior", "unexpected results"],
    "Group 4": ["expensive", "costly", "pricey"],
    "Group 5": ["time-consuming", "redundant", "complex", "bottleneck", "manual", "error-prone", "lack of integration", "unnecessary steps", "difficulty in finding information", "inconsistent processes", "inefficient"]
}

benefits = {
    "Group 1": ["easy", "simple", "user-friendly"],
    "Group 2": ["efficient", "streamlined", "productive", "time-efficient", "quick", "rapid", "expedited", "accelerated", "simplified", "automated", "expedited", "time-saving", "swift", "expedient", "efficiently", "speedy", "time-effective", "prompt", "time-optimized", "time-smart"],
    "Group 3": ["accurate", "reliable", "dependable"],
    "Group 4": ["economical", "budget-friendly", "cost-saving", "efficient", "value", "affordable", "roi", "thrifty", "practical", "money-saving", "resourceful", "frugal", "wise", "strategic", "practical", "reasonable", "cost-conscious", "smart", "judicious", "saver"],
    "Group 5": ["improved", "enhanced", "better"]
}

# Create counters for keyword occurrences
pains_count = {key: 0 for key in pains.keys()}
benefits_count = {key: 0 for key in benefits.keys()}

# Create sets to store sentences with keywords
pains_sentences = {key: set() for key in pains.keys()}
benefits_sentences = {key: set() for key in benefits.keys()}

# Loop through paragraphs in the .docx file
for paragraph in doc.paragraphs:
    text = paragraph.text.lower() # Convert text to lowercase for case-insensitive matching

    # Check for keywords in pains
    for group, keywords in pains.items():
        for keyword in keywords:
            if keyword in text:
                pains_count[group] += 1
                pains_sentences[group].add((text, keyword))

    # Check for keywords in benefits
    for group, keywords in benefits.items():
        for keyword in keywords:
            if keyword in text:
                benefits_count[group] += 1
                benefits_sentences[group].add((text, keyword))

# Create a new Excel workbook
wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Summary"

# Write headers for summary sheet
ws_summary['A1'] = "Keyword"
ws_summary['B1'] = "Count"
ws_summary['C1'] = "Group"
ws_summary['D1'] = "Set"

# Write data for summary sheet
row = 2
for group, keywords in pains.items():
    for keyword in keywords:
        ws_summary['A' + str(row)] = keyword
        ws_summary['B' + str(row)] = pains_count[group]
        ws_summary['C' + str(row)] = group
        ws_summary['D' + str(row)] = "Pains"
        row += 1

for group, keywords in benefits.items():
    for keyword in keywords:
        ws_summary['A' + str(row)] = keyword
        ws_summary['B' + str(row)] = benefits_count[group]
        ws_summary['C' + str(row)] = group
        ws_summary['D' + str(row)] = "Benefits"
        row += 1

# Create sheets for pains and benefits
ws_pains = wb.create_sheet("Pains")
ws_benefits = wb.create_sheet("Benefits")

# Write headers for pains and benefits sheets
ws_pains['A1'] = "Group"
ws_pains['B1'] = "Keyword"
ws_pains['C1'] = "Sentence"

ws_benefits['A1'] = "Group"
ws_benefits['B1'] = "Keyword"
ws_benefits['C1'] = "Sentence"

# Write data for pains and benefits sheets
for group, sentences in pains_sentences.items():
    row = 2
    for sentence, keyword in sentences:
        ws_pains['A' + str(row)] = group
        ws_pains['B' + str(row)] = keyword
        ws_pains['C' + str(row)] = sentence
        row += 1

for group, sentences in benefits_sentences.items():
    row = 2
    for sentence, keyword in sentences:
        ws_benefits['A' + str(row)] = group
        ws_benefits['B' + str(row)] = keyword
        ws_benefits['C' + str(row)] = sentence
        row += 1

# Auto-size columns for all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# Save the workbook to a file
wb.save("outputs.xlsx")

print("Excel file 'outputs.xlsx' has been created in the same folder.")
